import os
import sys
import io
import json
import logging
import requests
import pandas as pd
import webbrowser
import time
import socket
import sqlite3
from threading import Timer, Thread, Lock, RLock
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime, timedelta
from collections import OrderedDict
import math
import re
import concurrent.futures
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ─── EXE PATH HANDLING ───────────────────────────────────────────
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS                        
    EXE_DIR = os.path.dirname(sys.executable)      
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXE_DIR = BASE_DIR

template_dir = os.path.join(BASE_DIR, 'templates')
app = Flask(__name__, template_folder=template_dir)

# ─── FEATURE: SETTINGS MEMORY ────────────────────────────────────
SETTINGS_FILE = os.path.join(EXE_DIR, "settings.json")

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r') as f: return json.load(f)
        except Exception as e: logger.error(f"Error reading settings: {e}")
    return {"data_folder": "bhav_data"}

def save_settings(new_settings):
    settings = load_settings()
    settings.update(new_settings)
    try:
        with open(SETTINGS_FILE, 'w') as f: json.dump(settings, f, indent=4)
    except Exception as e: logger.error(f"Error saving settings: {e}")

# ─── FEATURE: SQLITE DATABASE SETUP (Point 3) ────────────────────
DB_PATH = os.path.join(EXE_DIR, "rangewise_market.db")

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS bhavdata (
            SYMBOL TEXT,
            DATE TEXT,
            OPEN_PRICE REAL,
            HIGH_PRICE REAL,
            LOW_PRICE REAL,
            CLOSE_PRICE REAL,
            PREV_CLOSE REAL,
            TTL_TRD_QNTY REAL,
            DELIV_PER REAL,
            UNIQUE(SYMBOL, DATE) ON CONFLICT REPLACE
        )
    ''')
    # Indexes for lightning-fast search
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_date ON bhavdata(DATE)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_symbol_date ON bhavdata(SYMBOL, DATE)')
    conn.commit()
    conn.close()

def get_db_summary():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT count(*) FROM bhavdata")
    total_records = cursor.fetchone()[0]
    cursor.execute("SELECT min(DATE), max(DATE) FROM bhavdata")
    date_range = cursor.fetchone()
    conn.close()
    return {"total_records": total_records, "min_date": date_range[0], "max_date": date_range[1]}

init_db() # App start hote hi DB ready ho jayega

def get_db_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn
# ────────────────────────────────────────────────────────────────

MAX_DAYS_SEARCH   = 10
REQUEST_TIMEOUT   = 30
DEFAULT_THRESHOLD = 500
DEFAULT_DELIV_PER = 40

# ─── HEARTBEAT (AUTO-SHUTDOWN) LOGIC ────────────────────────────
# 60s warmup buffer — browser open hone aur pehla ping aane tak ka time
LAST_PING_TIME = time.time() + 60

REQUEST_TIMEOUT = 15
MAX_DAYS_SEARCH = 30
DB_PATH = os.path.join(EXE_DIR, "rangewise_market.db")
LAST_PING_TIME = time.time()

@app.route("/api/ping", methods=["POST"])
def ping():
    global LAST_PING_TIME
    LAST_PING_TIME = time.time()
    return jsonify({"status": "alive"})

def monitor_heartbeat():
    global LAST_PING_TIME
    while True:
        time.sleep(30)  # Har 30s mein check karo — 3s zyada tha
        # 300s timeout: page reload + tab switch + browser throttling        # Ping check (Timeout increased to 1200s to avoid accidental shutdown)
        if time.time() - LAST_PING_TIME > 1200:
            logger.info("Heartbeat lost (300s timeout). Shutting down server...")
            os._exit(0)
# ────────────────────────────────────────────────────────────────

COLUMN_MAPPINGS = {
    'PREV_CLOSE':   ['PREV_CLOSE', 'PREVCLOSE', 'PREV CLOSE'],
    'OPEN_PRICE':   ['OPEN_PRICE', 'OPEN PRICE', 'OPEN'],
    'HIGH_PRICE':   ['HIGH_PRICE', 'HIGH PRICE', 'HIGH'],
    'LOW_PRICE':    ['LOW_PRICE',  'LOW PRICE',  'LOW'],
    'CLOSE_PRICE':  ['CLOSE_PRICE','CLOSE PRICE','CLOSE'],
    'TTL_TRD_QNTY': ['TTL_TRD_QNTY','TOTTRDQTY','TOTAL_TRADED_QUANTITY'],
    'DELIV_QTY':    ['DELIV_QTY',  'DELIVQTY',  'DELIVERABLE_QTY'],
    'DELIV_PER':    ['DELIV_PER',  'DELIVPER',  'DELY_QT_TO_TRD_QTY'],
    'NO_OF_TRADES': ['NO_OF_TRADES','NO OF TRADES','TRADES'],
}

class LRUCache(OrderedDict):
    def __init__(self, maxsize=50):
        super().__init__()
        self.maxsize = maxsize
        self._lock = RLock()   # RLock — same thread ek se zyada baar le sakta hai (reentrant)
    def __setitem__(self, key, value):
        with self._lock:
            if super().__contains__(key): self.move_to_end(key)  # super() call — deadlock avoid
            super().__setitem__(key, value)
            if len(self) > self.maxsize:
                oldest = next(iter(self))
                super().__delitem__(oldest)  # super() call — recursive lock avoid
    def __getitem__(self, key):
        with self._lock:
            value = super().__getitem__(key)
            self.move_to_end(key)
            return value
    def __contains__(self, key):
        with self._lock:
            return super().__contains__(key)

DATAFRAME_CACHE = LRUCache(maxsize=50)
API_CACHE_LOCK  = Lock()

_MARKET_HOLIDAYS      = set()
_SPECIAL_WORKING_DAYS = set()

def _load_holidays():
    global _MARKET_HOLIDAYS, _SPECIAL_WORKING_DAYS
    try:
        df = pd.read_excel(os.path.join(EXE_DIR, "MARKET_HOLIDAYS.xlsx"), sheet_name='Holidays')
        _MARKET_HOLIDAYS = set(pd.to_datetime(df['Date']).dt.strftime('%Y-%m-%d').tolist())
    except: pass
    
    try:
        df = pd.read_excel(os.path.join(EXE_DIR, "SPECIAL_WORKING_DAYS.xlsx"), sheet_name='Working Days')
        _SPECIAL_WORKING_DAYS = set(pd.to_datetime(df['Date']).dt.strftime('%Y-%m-%d').tolist())
    except: pass

_HOLIDAYS_LOADED = False
def ensure_holidays():
    global _HOLIDAYS_LOADED
    if not _HOLIDAYS_LOADED:
        _load_holidays()
        _HOLIDAYS_LOADED = True

def is_trading_day(d):
    ensure_holidays()
    s = d.strftime('%Y-%m-%d')
    if s in _SPECIAL_WORKING_DAYS: return True
    if d.weekday() >= 5:            return False
    if s in _MARKET_HOLIDAYS:       return False
    return True

def get_trading_dates(start_date, end_date):
    return [d.date() for d in pd.date_range(start=start_date, end=end_date, freq='D') if is_trading_day(d.date())]

def create_nse_session():
    session = requests.Session()
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://www.nseindia.com/",
        "Connection": "keep-alive",
        "Sec-Ch-Ua": '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
    }
    session.headers.update(headers)
    # Visit home page to get cookies
    try:
        session.get("https://www.nseindia.com", timeout=15)
        # Some NSE APIs need the cookies set from the search or get-quote page
        session.get("https://www.nseindia.com/market-data/live-equity-market", timeout=15)
    except Exception as e:
        logger.warning(f"Failed to initialize NSE session: {e}")
    return session

def nse_bhav_url_for(date_obj):
    fn  = f"sec_bhavdata_full_{date_obj.strftime('%d%m%Y')}.csv"
    url = f"https://nsearchives.nseindia.com/products/content/{fn}"
    return url, fn

def file_exists_locally(date_obj, data_folder):
    _, fn = nse_bhav_url_for(date_obj)
    fp = os.path.join(data_folder, fn)
    return fp if os.path.exists(fp) else None

def download_bhavcopy(date_obj, data_folder):
    os.makedirs(data_folder, exist_ok=True)
    existing = file_exists_locally(date_obj, data_folder)
    if existing: return existing
    url, fn = nse_bhav_url_for(date_obj)
    fp = os.path.join(data_folder, fn)
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Connection": "keep-alive",
        "Referer": "https://www.nseindia.com/"
    }

    for attempt in range(3):
        try:
            r = requests.get(url, headers=headers, timeout=REQUEST_TIMEOUT)
            if r.status_code == 200:
                with open(fp, "wb") as f: f.write(r.content)
                return fp
            elif r.status_code in [403, 401]:
                time.sleep(1.5 + attempt)
            elif r.status_code == 404:
                break
        except Exception:
            time.sleep(1)
    return None

def safe_get_col(df, col, suffix=""):
    full = f"{col}{suffix}"
    if full in df.columns: return df[full]
    for p in COLUMN_MAPPINGS.get(col, [col]):
        candidate = f"{p}{suffix}"
        if candidate in df.columns: return df[candidate]
    return pd.Series([pd.NA] * len(df), index=df.index)

# ─── THE NEW HYBRID DATA ENGINE (SQLite + CSV) ───────────────────
def get_data_hybrid(date_obj, data_folder):
    date_str = date_obj.strftime('%Y-%m-%d')

    # Level 1: Check Memory Cache
    if date_str in DATAFRAME_CACHE:
        return DATAFRAME_CACHE[date_str].copy()

    # Level 2: Check SQLite Database
    try:
        conn = sqlite3.connect(DB_PATH)
        query = "SELECT * FROM bhavdata WHERE DATE = ?"
        df_db = pd.read_sql(query, conn, params=(date_str,))
        conn.close()
        if not df_db.empty:
            DATAFRAME_CACHE[date_str] = df_db
            return df_db.copy()
    except Exception as e:
        logger.error(f"DB Read Error: {e}")

    # Level 3: Check Local CSV / Download from NSE
    fp = file_exists_locally(date_obj, data_folder)
    if not fp:
        fp = download_bhavcopy(date_obj, data_folder)

    if not fp: return None

    # Level 4: Process CSV and Push to SQLite
    try:
        df = pd.read_csv(fp)
        df.columns = [c.strip().upper().replace(' ', '_') for c in df.columns]

        if "SERIES" in df.columns:
            df["SERIES"] = df["SERIES"].astype(str).str.strip()
            df = df[df["SERIES"].isin(["EQ", "BE"])]

        # Standardize strictly for Database
        std_df = pd.DataFrame()
        std_df['SYMBOL'] = df['SYMBOL']
        std_df['DATE'] = date_str
        std_df['OPEN_PRICE'] = pd.to_numeric(safe_get_col(df, 'OPEN_PRICE'), errors='coerce')
        std_df['HIGH_PRICE'] = pd.to_numeric(safe_get_col(df, 'HIGH_PRICE'), errors='coerce')
        std_df['LOW_PRICE'] = pd.to_numeric(safe_get_col(df, 'LOW_PRICE'), errors='coerce')
        std_df['CLOSE_PRICE'] = pd.to_numeric(safe_get_col(df, 'CLOSE_PRICE'), errors='coerce')
        std_df['PREV_CLOSE'] = pd.to_numeric(safe_get_col(df, 'PREV_CLOSE'), errors='coerce')
        std_df['TTL_TRD_QNTY'] = pd.to_numeric(safe_get_col(df, 'TTL_TRD_QNTY'), errors='coerce')
        std_df['DELIV_PER'] = pd.to_numeric(safe_get_col(df, 'DELIV_PER'), errors='coerce')

        std_df.dropna(subset=['SYMBOL', 'CLOSE_PRICE'], inplace=True)

        # Save to SQLite DB — UNIQUE constraint ON CONFLICT REPLACE handle karega duplication
        conn = sqlite3.connect(DB_PATH)
        std_df.to_sql('bhavdata', conn, if_exists='append', index=False)
        conn.commit()
        conn.close()

        DATAFRAME_CACHE[date_str] = std_df
        return std_df.copy()
    except Exception as e:
        logger.error(f"Hybrid Process Error: {e}")
        return None
# ────────────────────────────────────────────────────────────────

# ─── UPDATED DATA FINDERS (Using Hybrid Engine) ─────────────────
def find_df_before(date_obj, data_folder, max_days=30):
    current = date_obj.date() if isinstance(date_obj, datetime) else date_obj
    days_checked = 0
    while days_checked < max_days:
        if is_trading_day(current):
            df = get_data_hybrid(current, data_folder)
            if df is not None and not df.empty:
                return current, df
        current -= timedelta(days=1)
        days_checked += 1
    return None

def find_df_after(date_obj, data_folder, max_days=5):
    today = datetime.now().date()
    cur   = date_obj + timedelta(days=1)
    cur   = cur.date() if isinstance(cur, datetime) else cur
    for _ in range(max_days):
        if cur > today: return None
        if is_trading_day(cur):
            df = get_data_hybrid(cur, data_folder)
            if df is not None and not df.empty: return (cur, df)
        cur += timedelta(days=1)
    return None

def get_next_n_trading_dfs(start_date, data_folder, n=7):
    result, today = [], datetime.now().date()
    cur = start_date + timedelta(days=1)
    cur = cur.date() if isinstance(cur, datetime) else cur
    attempts = 0
    while len(result) < n and attempts < n * 3:
        if cur > today: break
        if is_trading_day(cur):
            df = get_data_hybrid(cur, data_folder)
            if df is not None and not df.empty: result.append((cur, df))
        cur += timedelta(days=1)
        attempts += 1
    return result
# ────────────────────────────────────────────────────────────────

def safe_pct(num, den):
    num = pd.to_numeric(num, errors='coerce')
    den = pd.to_numeric(den, errors='coerce')
    result = (num / den.replace(0, pd.NA) * 100).round(2)
    return result.replace([float('inf'), float('-inf')], pd.NA)

def get_7day_status(g):
    if pd.isna(g) or g < 0.10: return 'NO TRADE'
    if g >= 1:                 return 'WIN'
    return 'LOSS'

def load_index_constituents():
    try:
        path = os.path.join(BASE_DIR, "indices.json")
        with open(path) as f: return {k: set(v) for k, v in json.load(f).items()}
    except Exception: return {}

def load_thresholds_json():
    try:
        path = os.path.join(BASE_DIR, "Thresholds.json")
        with open(path) as f:
            data = json.load(f)
        return {item.get("Symbol"): {"threshold": item.get("Threshold", DEFAULT_THRESHOLD)} for item in data if item.get("Symbol")}
    except Exception: return {}

INDICES    = load_index_constituents()
THRESHOLDS = load_thresholds_json()
_load_holidays()

def process_single_date(signal_date, threshold_mode, custom_threshold, deliv_per_threshold, data_folder, enable_next_day=True, enable_7day=True):
    try:
        # Ab seedha DataFrame milega hybrid engine se! (Code is much cleaner)
        cur_found = find_df_before(signal_date, data_folder)
        if not cur_found: return None, None
        cur_date, df_cur = cur_found

        prev_found = find_df_before(cur_date - timedelta(days=1), data_folder)
        if not prev_found: return None, None
        prev_date, df_prev = prev_found

        df_merge = pd.merge(df_prev, df_cur, on="SYMBOL", suffixes=("_PREV","_CURR"), how="inner")
        qty_prev = df_merge["TTL_TRD_QNTY_PREV"]
        qty_curr = df_merge["TTL_TRD_QNTY_CURR"]
        df_merge["QTY_GROWTH_%"] = ((qty_curr - qty_prev) / qty_prev.replace({0: pd.NA}) * 100).round(2)
        deliv_curr = df_merge["DELIV_PER_CURR"]

        df_sig = pd.DataFrame({
            "SYMBOL":       df_merge["SYMBOL"].values,
            "QTY_GROWTH_%": df_merge["QTY_GROWTH_%"].values,
            "DELIV_PER":    deliv_curr.values,
        })
        
        if threshold_mode == "Use Thresholds.json" and THRESHOLDS:
            df_sig = df_sig[df_sig["SYMBOL"].isin(THRESHOLDS.keys())].copy()
            df_sig["Threshold Used"]   = df_sig["SYMBOL"].map(
                {s: d["threshold"] for s, d in THRESHOLDS.items()}
            ).fillna(DEFAULT_THRESHOLD)
            df_sig["Threshold Source"] = "Custom (JSON)"
        else:
            df_sig["Threshold Used"]   = float(custom_threshold)
            df_sig["Threshold Source"] = "Default (User-Set)"

        df_sig["QTY_GROWTH_%"]  = pd.to_numeric(df_sig["QTY_GROWTH_%"],  errors="coerce")
        df_sig["Threshold Used"]= pd.to_numeric(df_sig["Threshold Used"], errors="coerce")
        df_sig["DELIV_PER"]     = pd.to_numeric(df_sig["DELIV_PER"],      errors="coerce")

        df_sig = df_sig[
            (df_sig["QTY_GROWTH_%"]   > df_sig["Threshold Used"]) &
            (df_sig["DELIV_PER"]      >= deliv_per_threshold)
        ].copy()

        if df_sig.empty: return None, None

        df_cs = df_cur[df_cur["SYMBOL"].isin(df_sig["SYMBOL"].tolist())].copy()
        df_cs = df_cs.merge(df_sig[["SYMBOL", "QTY_GROWTH_%", "Threshold Used", "Threshold Source"]], on="SYMBOL", how="left")

        sig_res = pd.DataFrame({
            "SIGNAL_DATE":                cur_date.strftime("%d-%b-%Y"),
            "SYMBOL":                     df_cs["SYMBOL"].values,
            "QTY_GROWTH_%":               df_cs["QTY_GROWTH_%"].values,
            "Threshold Used":             df_cs["Threshold Used"].values,
            "Threshold Source":           df_cs["Threshold Source"].values,
            "PREV_CLOSE":                 df_cs["PREV_CLOSE"].values,
            "OPEN":                       df_cs["OPEN_PRICE"].values,
            "HIGH":                       df_cs["HIGH_PRICE"].values,
            "LOW":                        df_cs["LOW_PRICE"].values,
            "CLOSE":                      df_cs["CLOSE_PRICE"].values,
            "% Change":                   safe_pct(df_cs["CLOSE_PRICE"] - df_cs["OPEN_PRICE"], df_cs["OPEN_PRICE"]).values,
            "% GROWTH (OPEN→HIGH)":  safe_pct(df_cs["HIGH_PRICE"] - df_cs["OPEN_PRICE"], df_cs["OPEN_PRICE"]).values,
            "% GROWTH (OPEN→CLOSE)": safe_pct(df_cs["CLOSE_PRICE"] - df_cs["OPEN_PRICE"], df_cs["OPEN_PRICE"]).values,
            "DELIV_PER":                  df_cs["DELIV_PER"].values,
        })

        sig_res["STATUS"] = "LOSS"
        sig_res.loc[sig_res["% GROWTH (OPEN→HIGH)"] >= 1,  "STATUS"] = "WIN"
        sig_res.loc[
            sig_res["% GROWTH (OPEN→HIGH)"].isna() |
            (sig_res["% GROWTH (OPEN→HIGH)"] < 0.10), "STATUS"
        ] = "NO TRADE"

        nd_res = None
        if enable_next_day:
            nf = find_df_after(cur_date, data_folder)
            if nf:
                nd, df_nd_raw = nf
                df_nd_raw = df_nd_raw[df_nd_raw["SYMBOL"].isin(df_sig["SYMBOL"].tolist())].copy()
                if not df_nd_raw.empty:
                    # sig_res se signal-day columns lete hain — rename karo taaki df_nd_raw ke
                    # PREV_CLOSE/OPEN/HIGH/LOW/CLOSE se conflict na ho (_x/_y se bachna)
                    sig_for_merge = sig_res[[
                        "SYMBOL", "SIGNAL_DATE", "QTY_GROWTH_%",
                        "PREV_CLOSE", "OPEN", "HIGH", "LOW", "CLOSE", "% Change"
                    ]].rename(columns={
                        "PREV_CLOSE": "_s_prev", "OPEN": "_s_open",
                        "HIGH": "_s_high", "LOW": "_s_low",
                        "CLOSE": "_s_close", "% Change": "_s_pct"
                    })
                    df_nd_raw = df_nd_raw.merge(sig_for_merge, on="SYMBOL", how="inner")
                    df_nd_raw = df_nd_raw.merge(
                        df_sig[["SYMBOL", "Threshold Used", "Threshold Source"]],
                        on="SYMBOL", how="left"
                    )

                    on_ = df_nd_raw["OPEN_PRICE"]
                    hn_ = df_nd_raw["HIGH_PRICE"]
                    cn_ = df_nd_raw["CLOSE_PRICE"]
                    ln_ = df_nd_raw["LOW_PRICE"]

                    nd_res = pd.DataFrame({
                        "SIGNAL_DATE":                df_nd_raw["SIGNAL_DATE"].values,
                        "SYMBOL":                     df_nd_raw["SYMBOL"].values,
                        "QTY_GROWTH_%":               df_nd_raw["QTY_GROWTH_%"].values,
                        "PREV_CLOSE":                 df_nd_raw["_s_prev"].values,
                        "OPEN":                       df_nd_raw["_s_open"].values,
                        "HIGH":                       df_nd_raw["_s_high"].values,
                        "LOW":                        df_nd_raw["_s_low"].values,
                        "CLOSE":                      df_nd_raw["_s_close"].values,
                        "% Change":                   df_nd_raw["_s_pct"].values,
                        "NEXT_DAY_DATE":              nd.strftime("%d-%b-%Y"),
                        "NEXT_DAY_OPEN":              on_.values,
                        "NEXT_DAY_HIGH":              hn_.values,
                        "NEXT_DAY_LOW":               ln_.values,
                        "NEXT_DAY_CLOSE":             cn_.values,
                        "NEXT_DAY_% Change":          safe_pct(cn_ - on_, on_).values,
                        "% GROWTH (OPEN→HIGH)":  safe_pct(hn_ - on_, on_).values,
                        "% GROWTH (OPEN→CLOSE)": safe_pct(cn_ - on_, on_).values,
                        "DELIV_PER":                  df_nd_raw["DELIV_PER"].values,
                    })
                    nd_res["STATUS"] = "LOSS"
                    nd_res.loc[nd_res["% GROWTH (OPEN→HIGH)"] >= 1, "STATUS"] = "WIN"
                    nd_res.loc[
                        nd_res["% GROWTH (OPEN→HIGH)"].isna() |
                        (nd_res["% GROWTH (OPEN→HIGH)"] < 0.10), "STATUS"
                    ] = "NO TRADE"

        if enable_7day:
            trading_days = get_next_n_trading_dfs(cur_date, data_folder)
            if trading_days:
                syms = sig_res["SYMBOL"].tolist()
                s7 = pd.DataFrame({"SYMBOL": syms})

                d1_date, df_d1 = trading_days[0]
                df_d1 = df_d1[df_d1["SYMBOL"].isin(syms)].copy()
                s7 = s7.merge(pd.DataFrame({"SYMBOL": df_d1["SYMBOL"].values, "DAY1_OPEN": df_d1["OPEN_PRICE"].values}), on="SYMBOL", how="left")

                day_cols = []
                for di, (td, df_day) in enumerate(trading_days, 1):
                    try:
                        df_day = df_day[df_day["SYMBOL"].isin(syms)].copy()
                        tmp = pd.DataFrame({"SYMBOL": df_day["SYMBOL"].values, f"Day{di}_High": df_day["HIGH_PRICE"].values})
                        s7 = s7.merge(tmp, on="SYMBOL", how="left")
                        s7[f"Day{di}_%"] = ((s7[f"Day{di}_High"] - s7["DAY1_OPEN"]) / s7["DAY1_OPEN"] * 100).round(2)
                        day_cols.append(f"Day{di}_%")
                    except Exception:
                        continue

                if day_cols:
                    s7["% 7 DAYS GROWTH (OPEN→HIGH)"] = s7[day_cols].max(axis=1).round(2)
                else:
                    s7["% 7 DAYS GROWTH (OPEN→HIGH)"] = 0.0

                sig_res = sig_res.merge(s7[["SYMBOL", "% 7 DAYS GROWTH (OPEN→HIGH)"]], on="SYMBOL", how="left")
                if nd_res is not None:
                    nd_res = nd_res.merge(s7[["SYMBOL", "% 7 DAYS GROWTH (OPEN→HIGH)"]], on="SYMBOL", how="left")

        return sig_res, nd_res

    except Exception as e:
        logger.error(f"process_single_date error: {e}", exc_info=True)
        return None, None

def auto_adjust_excel_width(writer, df, sheet_name):
    ws  = writer.sheets[sheet_name]
    hf  = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    hft = Font(bold=True, color="FFFFFF", size=11)
    for i, col in enumerate(df.columns, 1):
        cl = get_column_letter(i)
        mx = max(int(df[col].astype(str).apply(len).max() or 0), len(str(col)))
        ws.column_dimensions[cl].width = min(mx + 2, 30)
        hc = ws[f"{cl}1"]
        hc.fill, hc.font = hf, hft
        hc.alignment = Alignment(horizontal="center", vertical="center")

def _run_analysis(data):
    start_date       = datetime.strptime(data["start_date"], "%Y-%m-%d").date()
    end_date         = datetime.strptime(data["end_date"],   "%Y-%m-%d").date()
    threshold_mode   = data.get("threshold_mode", "Use Thresholds.json")
    custom_threshold = float(data.get("custom_threshold", DEFAULT_THRESHOLD))
    deliv_per        = float(data.get("deliv_per", DEFAULT_DELIV_PER))
    enable_7day      = data.get("enable_7day", True)
    remove_negative  = data.get("remove_negative", False)
    index_filter     = data.get("index_filter", "All")
    
    settings = load_settings()
    data_folder      = data.get("data_folder", "").strip()
    if not data_folder: data_folder = settings.get("data_folder", "bhav_data")
    save_settings({"data_folder": data_folder})

    if start_date > end_date: return {"error": "Start date must be before end date"}

    trading_dates = get_trading_dates(start_date, end_date)
    if not trading_dates: return {"error": "No trading days found in selected range"}

    all_sig, all_nd = [], []
    for td in trading_dates:
        sr, nr = process_single_date(td, threshold_mode, custom_threshold,
                                     deliv_per, data_folder, True, enable_7day)
        if sr is not None and not sr.empty: all_sig.append(sr)
        if nr is not None and not nr.empty: all_nd.append(nr)

    if not all_sig: return {"error": "No signals found in selected date range"}

    df_combined = pd.concat(all_sig, ignore_index=True)
    df_nd_all   = pd.concat(all_nd,  ignore_index=True) if all_nd else pd.DataFrame()

    if index_filter != "All" and index_filter in INDICES:
        syms = INDICES[index_filter]
        df_combined = df_combined[df_combined["SYMBOL"].isin(syms)].copy()
        if not df_nd_all.empty:
            df_nd_all = df_nd_all[df_nd_all["SYMBOL"].isin(syms)].copy()

    if remove_negative and "% Change" in df_combined.columns:
        df_combined = df_combined[pd.to_numeric(df_combined["% Change"], errors='coerce') >= 0].copy()
        if not df_nd_all.empty and "% Change" in df_nd_all.columns:
            df_nd_all = df_nd_all[pd.to_numeric(df_nd_all["% Change"], errors='coerce') >= 0].copy()

    if df_combined.empty: return {"error": "No data matches current filter criteria"}

    if not df_nd_all.empty:
        if 'STATUS' in df_nd_all.columns:
            df_nd_all.rename(columns={'STATUS': 'STATUS (1D)'}, inplace=True)
        if '% 7 DAYS GROWTH (OPEN→HIGH)' in df_nd_all.columns:
            df_nd_all['STATUS (7D)'] = df_nd_all['% 7 DAYS GROWTH (OPEN→HIGH)'].apply(get_7day_status)
            
            mask_no_trade = df_nd_all['STATUS (1D)'] == 'NO TRADE'
            df_nd_all.loc[mask_no_trade, 'STATUS (7D)'] = 'NO TRADE'

    has_7d = 'STATUS (7D)' in df_nd_all.columns if not df_nd_all.empty else False

    def df_to_records(df):
        df2 = df.copy()
        for c in df2.select_dtypes(include='float').columns:
            df2[c] = df2[c].round(2)
        records = df2.to_dict(orient='records')
        return [{k: (None if pd.isna(v) else v) for k, v in r.items()} for r in records]

    def build_breakdown(df_src, status_col):
        if df_src.empty or status_col not in df_src.columns: return []
        grp = df_src.groupby('SIGNAL_DATE')
        bd  = grp.agg({'SYMBOL': 'count'}).rename(columns={'SYMBOL': 'TOTAL'})
        bd['WIN']       = grp[status_col].apply(lambda x: (x == 'WIN').sum())
        bd['LOSS']      = grp[status_col].apply(lambda x: (x == 'LOSS').sum())
        bd['NO TRADE']  = grp[status_col].apply(lambda x: (x == 'NO TRADE').sum())
        bd['NET VALUE'] = bd['TOTAL'] - bd['NO TRADE']
        
        bd['WIN %'] = 0.0
        mask = bd['NET VALUE'] > 0
        bd.loc[mask, 'WIN %'] = (bd.loc[mask, 'WIN'] / bd.loc[mask, 'NET VALUE'] * 100).round(2)
        
        records = bd.reset_index().to_dict(orient='records')
        return [{k: (None if pd.isna(v) else v) for k, v in r.items()} for r in records]

    nd_records = []
    kpi        = {}
    if not df_nd_all.empty:
        nd_records = df_to_records(df_nd_all.sort_values(['SIGNAL_DATE', 'SYMBOL']))
        total  = len(df_nd_all)
        s1d    = df_nd_all.get('STATUS (1D)', pd.Series(dtype=str))
        win1   = int((s1d == 'WIN').sum())
        loss1  = int((s1d == 'LOSS').sum())
        notrd  = int((s1d == 'NO TRADE').sum())
        avg_r  = df_nd_all.get('% GROWTH (OPEN→HIGH)', pd.Series(dtype=float)).mean()
        win7   = int((df_nd_all.get('STATUS (7D)', pd.Series(dtype=str)) == 'WIN').sum())  if has_7d else 0
        loss7  = int((df_nd_all.get('STATUS (7D)', pd.Series(dtype=str)) == 'LOSS').sum()) if has_7d else 0
        
        net_trades = total - notrd
        
        kpi = {
            "total":   total,
            "win1d":   win1,  
            "win1d_pct":  round(win1 / net_trades * 100, 1) if net_trades > 0 else 0,
            "loss1d":  loss1,
            "notrade": notrd,
            "avg_ret": round(float(avg_r), 2) if pd.notna(avg_r) else 0,
            "win7d":   win7,  
            "win7d_pct":  round(win7 / net_trades * 100, 1) if net_trades > 0 else 0,
            "loss7d":  loss7,
        }
    else:
        # Provide signal day records, but blank out the next-day (result) columns
        df_display = df_combined.copy()
        df_display['STATUS (1D)'] = pd.NA
        df_display['% GROWTH (OPEN→HIGH)'] = pd.NA
        df_display['% GROWTH (OPEN→CLOSE)'] = pd.NA
        nd_records = df_to_records(df_display.sort_values(['SIGNAL_DATE', 'SYMBOL']))

    return {
        "kpi":          kpi,
        "nd_records":   nd_records,
        "breakdown_1d": build_breakdown(df_nd_all, 'STATUS (1D)'),
        "breakdown_7d": build_breakdown(df_nd_all, 'STATUS (7D)') if has_7d else [],
        "signal_count": len(df_combined),
        "has_7d":       has_7d,
        "params":       data,
    }

# ─── EXTRA API LOGIC FOR FORECAST & ANALYZER ──────────────────────
REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.5",
    "Referer": "https://www.nseindia.com/",
    "Connection": "keep-alive"
}

def create_nse_session(timeout=15):
    session = requests.Session()
    session.headers.update(REQUEST_HEADERS)
    try:
        session.get("https://www.nseindia.com", timeout=timeout)
        session.get("https://www.nseindia.com/market-data/live-equity-market", timeout=timeout)
    except Exception: pass
    return session

def rank_from_nearlow(x):
    if pd.isna(x): return "Rank Unknown"
    x = float(x)
    for i in range(1, 8):
        if x < i * 10: return f"Rank {i} ({(i-1)*10}-{i*10}%)"
    return "Rank 8 (70-100%)"

API_CACHE = {}

def fetch_nifty_index(index_name):
    cache_key = f"nifty_{index_name}"
    now = time.time()
    with API_CACHE_LOCK:
        if cache_key in API_CACHE and (now - API_CACHE[cache_key]['time'] < 180):
            return API_CACHE[cache_key]['data']

    session = create_nse_session()
    link = f"https://www.nseindia.com/api/equity-stockIndices?index={index_name.replace(' ', '%20')}"
    resp = session.get(link, timeout=15)
    resp.raise_for_status()
    data = resp.json()
    df = pd.DataFrame(data["data"])
    
    for col in ["lastPrice", "yearHigh", "yearLow"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '').str.strip(), errors='coerce')
    
    if all(c in df.columns for c in ["lastPrice", "yearHigh", "yearLow"]):
        denom = (df["yearHigh"] - df["yearLow"]).replace(0, pd.NA)
        df["Near to low"] = ((df["lastPrice"] - df["yearLow"]) * 100 / denom).round(2)
        df["Profit %"] = ((df["yearHigh"] - df["lastPrice"]) / df["lastPrice"] * 100).round(2)
        df["Rank"] = df["Near to low"].apply(rank_from_nearlow)
    
    # Fill missing Company Name if empty
    if "Company Name" not in df.columns:
        df["Company Name"] = df["symbol"]
    
    result = df.fillna("")[["symbol", "Company Name", "lastPrice", "yearHigh", "yearLow", "Near to low", "Profit %", "Rank"]].to_dict('records')
    with API_CACHE_LOCK:
        API_CACHE[cache_key] = {'data': result, 'time': now}
    return result

def download_nse_etf_clean():
    session = create_nse_session()
    # NSE ETF CSV download link
    r = session.get("https://www.nseindia.com/api/etf?csv=true", timeout=15)
    r.raise_for_status()
    
    df = pd.read_csv(io.BytesIO(r.content))
    df.columns = [c.strip().upper() for c in df.columns]
    
    # Robust name detection
    possible_name_cols = ['NAME OF ETF', 'NAME OF INTERIM ETF', 'COMPANY NAME', 'NAME']
    name_col = next((c for c in possible_name_cols if c in df.columns), None)
    
    if not name_col:
        # Fallback to finding the first column that looks like a name
        name_col = df.columns[0]
        logger.warning(f"Could not find exact ETF name column. Falling back to: {name_col}")
        
    df = df.set_index(name_col)
    # Map common columns to consistent names
    col_map = {
        'SYMBOL': 'symbol',
        'LTP': 'lastPrice',
        '52W H': 'yearHigh',
        '52W L': 'yearLow',
    }
    # Update col_map with variations seen in NSE data
    for c in df.columns:
        cu = c.upper()
        if 'SYMBOL' in cu: col_map[c] = 'symbol'
        if 'LTP' in cu or 'LAST PRICE' in cu: col_map[c] = 'lastPrice'
        if '52W HIGH' in cu or '52WH' in cu: col_map[c] = 'yearHigh'
        if '52W LOW' in cu or '52WL' in cu: col_map[c] = 'yearLow'

    df = df.rename(columns=col_map)
    
    # Calculate essential metrics if missing
    if 'Near to low' not in df.columns and 'lastPrice' in df.columns and 'yearLow' in df.columns:
        df['lastPrice'] = pd.to_numeric(df['lastPrice'], errors='coerce')
        df['yearLow'] = pd.to_numeric(df['yearLow'], errors='coerce')
        df['Near to low'] = ((df['lastPrice'] - df['yearLow']) / df['yearLow']) * 100
        
    if 'Profit %' not in df.columns and 'lastPrice' in df.columns and 'yearHigh' in df.columns:
        df['yearHigh'] = pd.to_numeric(df['yearHigh'], errors='coerce')
        df['Profit %'] = ((df['yearHigh'] - df['lastPrice']) / df['lastPrice']) * 100

    def get_rank(near):
        if pd.isna(near): return "Unknown"
        if near <= 10: return "Rank 1 (0-10%)"
        if near <= 20: return "Rank 2 (10-20%)"
        if near <= 30: return "Rank 3 (20-30%)"
        if near <= 40: return "Rank 4 (30-40%)"
        if near <= 50: return "Rank 5 (40-50%)"
        if near <= 60: return "Rank 6 (50-60%)"
        if near <= 70: return "Rank 7 (60-70%)"
        return "Rank 8 (70-100%)"

    if 'Near to low' in df.columns:
        df['Rank'] = df['Near to low'].apply(get_rank)
    
    return df.reset_index().rename(columns={name_col: 'NAME OF ETF'})

@app.route("/")
def home():
    return render_template("home.html")

@app.route("/rangewise")
def rangewise():
    settings = load_settings()
    saved_path = settings.get("data_folder", "bhav_data")
    return render_template("rangewise.html",
                           indices=list(INDICES.keys()),
                           thresholds_count=len(THRESHOLDS),
                           default_threshold=DEFAULT_THRESHOLD,
                           default_deliv=DEFAULT_DELIV_PER,
                           saved_data_folder=saved_path)

@app.route("/forecast")
def forecast():
    return render_template("forecast.html")

@app.route("/analyzer")
def analyzer():
    settings = load_settings()
    saved_path = settings.get("data_folder", "bhav_data")
    return render_template("analyzer.html", indices=list(INDICES.keys()), saved_data_folder=saved_path)

@app.route("/db_info")
def db_info():
    try:
        # Removed get_db_summary() to speed up page landing
        return render_template("db_info.html", info=None)
    except Exception as e:
        logger.error(f"Error in /db_info: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/db/check_missing", methods=["POST"])
def api_db_check_missing():
    try:
        data = request.json
        start_str = data.get("start_date")
        end_str = data.get("end_date")
        
        if not start_str or not end_str:
            return jsonify({"error": "Start and End dates are required"}), 400
            
        start_date = pd.to_datetime(start_str).date()
        end_date = pd.to_datetime(end_str).date()
        
        # Get all intended trading dates in range
        intended_dates = []
        curr = start_date
        while curr <= end_date:
            if is_trading_day(curr):
                intended_dates.append(curr.strftime('%Y-%m-%d'))
            curr += timedelta(days=1)
            
        # Get actual dates in DB
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT DATE FROM bhavdata WHERE DATE BETWEEN ? AND ?", (start_str, end_str))
        existing_dates = {row[0] for row in cursor.fetchall()}
        conn.close()
        
        missing_dates = [d for d in intended_dates if d not in existing_dates]
        
        return jsonify({
            "success": True,
            "missing_dates": missing_dates,
            "total_intended": len(intended_dates),
            "total_missing": len(missing_dates)
        })
    except Exception as e:
        logger.error(f"Error in check_missing: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/db/summary")
def api_db_summary():
    try:
        summary = get_db_summary()
        return jsonify(summary)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/forecast/nifty", methods=["POST"])
def api_forecast_nifty():
    try:
        index_name = request.json.get("index", "NIFTY 50")
        data = fetch_nifty_index(index_name)
        return jsonify({"data": data})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/forecast/etf", methods=["GET"])
def api_forecast_etf():
    try:
        df = download_nse_etf_clean()
        return jsonify({"success": True, "data": df.to_dict(orient='records')})
    except Exception as e:
        logger.error(f"ETF Fetch Error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/forecast/export", methods=["POST"])
def api_forecast_export():
    try:
        data = request.json.get("data", [])
        filename = request.json.get("filename", "export")
        df = pd.DataFrame(data)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=filename[:30])
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f"{filename}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/analyzer/run", methods=["POST"])
def api_analyzer_run():
    try:
        data = request.json
        params = {
            "start_date": data["date"],
            "end_date":   data["date"],
            "threshold_mode":   data.get("threshold_mode", "Use Thresholds.json"),
            "custom_threshold": data.get("custom_threshold", 500),
            "deliv_per":        data.get("deliv_per", 40),
            "index_filter":     data.get("index_filter", "All"),
            "enable_7day":      data.get("enable_7day", True),
            "remove_negative":  data.get("remove_negative", False),
            "data_folder":      data.get("data_folder", "")
        }
        res = _run_analysis(params)
        if "error" in res:
            logger.warning(f"Analyzer error: {res['error']}")
            return jsonify(res), 400

        # nd_records = next-day data (tabhi milega jab agle din ka data available ho)
        # Agar nd_records empty ho (e.g. aaj ki date) toh signal_date records return karo
        records = res.get("nd_records", [])
        return jsonify({
            "data":         records,
            "kpi":          res.get("kpi", {}),
            "has_7d":       res.get("has_7d", False),
            "signal_count": res.get("signal_count", 0),
            "has_nd":       len(records) > 0
        })
    except Exception as e:
        logger.error(f"/api/analyzer/run error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/analyzer/export", methods=["POST"])
def api_analyzer_export():
    try:
        data = request.json.get("data", [])
        if not data:
            return jsonify({"error": "No data to export"}), 400
        df = pd.DataFrame(data)
        
        # Keep specific columns
        cols_to_keep = [
            "SYMBOL", "SIGNAL_DATE", "PREV_CLOSE", "OPEN", "HIGH", "LOW", "CLOSE", "% Change",
            "QTY_GROWTH_%", "DELIV_PER", "% GROWTH (OPEN→HIGH)", "% GROWTH (OPEN→CLOSE)", "STATUS (1D)"
        ]
        if "% 7 DAYS GROWTH (OPEN→HIGH)" in df.columns:
            cols_to_keep.extend(["% 7 DAYS GROWTH (OPEN→HIGH)", "STATUS (7D)"])
            
        cols_final = [c for c in cols_to_keep if c in df.columns]
        if cols_final: df = df[cols_final]
        
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Analyzer Results")
            auto_adjust_excel_width(writer, df, "Analyzer Results")
            
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name="analyzer_export.xlsx")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def clean_floats(d):
    if isinstance(d, dict): return {k: clean_floats(v) for k, v in d.items()}
    if isinstance(d, list): return [clean_floats(v) for v in d]
    if isinstance(d, float) and (math.isnan(d) or math.isinf(d)): return None
    return d

def format_df(df):
    df = df.set_index(df.columns[0])
    df.index = df.index.str.replace(r'\s*\+$', '', regex=True).str.strip()
    # Optional: Clean up index labels
    df.index = [str(x).replace("Sales +", "Sales").replace("Net Profit +", "Net Profit").strip() for x in df.index]
    return {"headers": list(df.columns), "data": clean_floats(df.to_dict(orient="index"))}

def fetch_symbol_fundamentals(symbol):
    url = f"https://www.screener.in/company/{symbol}/"
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        res = requests.get(url, headers=headers, timeout=10)
        if res.status_code != 200:
            return {"symbol": symbol, "success": False, "error": f"Failed with status {res.status_code}"}
        
        html_str = res.content.decode('utf-8', errors='ignore')
        from lxml import html
        tree = html.fromstring(res.content)
        sections = tree.xpath('//section')
        data = {}
        
        # Regex for pros and cons
        pros, cons = [], []
        pros_match = re.search(r'<div class="pros"[\s\S]*?<ul[^>]*>([\s\S]*?)</ul>', html_str)
        if pros_match:
            pros_items = re.findall(r'<li[^>]*>(.*?)</li>', pros_match.group(1), re.IGNORECASE | re.DOTALL)
            pros = [re.sub(r'<[^>]+>', '', p).strip() for p in pros_items]
            
        cons_match = re.search(r'<div class="cons"[\s\S]*?<ul[^>]*>([\s\S]*?)</ul>', html_str)
        if cons_match:
            cons_items = re.findall(r'<li[^>]*>(.*?)</li>', cons_match.group(1), re.IGNORECASE | re.DOTALL)
            cons = [re.sub(r'<[^>]+>', '', c).strip() for c in cons_items]
        data['analysis'] = {"pros": pros, "cons": cons}
        
        for sec in sections:
            h2 = sec.xpath('.//h2')
            table = sec.xpath('.//table')
            if h2 and table:
                title = h2[0].text_content().strip().lower()
                table_html = html.tostring(table[0])
                try:
                    df = pd.read_html(table_html)[0]
                    if 'profit & loss' in title:
                        data["profitLoss"] = format_df(df)
                    elif 'quarter' in title:
                        data["quarters"] = format_df(df)
                    elif 'balance sheet' in title:
                        data["balanceSheet"] = format_df(df)
                    elif 'cash flow' in title:
                        data["cashFlow"] = format_df(df)
                    elif 'ratios' in title:
                        data["ratios"] = format_df(df)
                    elif 'shareholding' in title:
                        data["shareholding"] = format_df(df)
                except: pass
        
        return {"symbol": symbol, "success": True, "data": data}
    except Exception as e:
        return {"symbol": symbol, "success": False, "error": str(e)}

@app.route("/api/fundamentals/batch", methods=["POST"])
def api_fundamentals_batch():
    try:
        symbols = request.json.get("symbols", [])
        if not symbols:
            return jsonify({"error": "No symbols provided"}), 400

        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            results = list(executor.map(fetch_symbol_fundamentals, symbols))

        return jsonify({"results": results})
    except Exception as e:
        logger.error(f"/api/fundamentals/batch error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/chart_data/<symbol>", methods=["GET"])
def api_chart_data(symbol):
    try:
        conn = get_db_conn()
        q = """
        SELECT DATE, OPEN_PRICE, HIGH_PRICE, LOW_PRICE, CLOSE_PRICE, TTL_TRD_QNTY 
        FROM bhavdata 
        WHERE SYMBOL = ? 
        ORDER BY DATE DESC 
        LIMIT 90
        """
        df = pd.read_sql_query(q, conn, params=(symbol,))
        conn.close()
        
        if df.empty:
            return jsonify([])
            
        # Reverse to ascending for the chart
        df = df.sort_values("DATE", ascending=True)
        
        chart_data = []
        for _, row in df.iterrows():
            chart_data.append({
                "time": row["DATE"],
                "open": float(row["OPEN_PRICE"]),
                "high": float(row["HIGH_PRICE"]),
                "low": float(row["LOW_PRICE"]),
                "close": float(row["CLOSE_PRICE"]),
                "value": float(row["TTL_TRD_QNTY"]) # value is used for volume histogram
            })
        return jsonify(chart_data)
    except Exception as e:
        logger.error(f"/api/chart_data error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/run", methods=["POST"])
def run_analysis():
    try:
        result = _run_analysis(request.json)
        if "error" in result:
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        logger.error(f"/api/run error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/export", methods=["POST"])
def export_excel():
    try:
        result = _run_analysis(request.json)
        if "error" in result: return jsonify(result), 400

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            nd = pd.DataFrame(result["nd_records"])
            if not nd.empty:
                nd.to_excel(writer, index=False, sheet_name="Next Day Analysis")
                auto_adjust_excel_width(writer, nd, "Next Day Analysis")
            for key, sname in [("breakdown_1d", "Breakdown 1D"), ("breakdown_7d", "Breakdown 7D")]:
                bd_data = result.get(key, [])
                if bd_data:
                    df_bd = pd.DataFrame(bd_data)
                    df_bd['WIN %'] = df_bd['WIN %'].map(lambda x: f"{float(x):.2f}%" if pd.notnull(x) else x)
                    df_bd.to_excel(writer, index=False, sheet_name=sname)
                    auto_adjust_excel_width(writer, df_bd, sname)

        buf.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(buf, as_attachment=True,
                         download_name=f"rangewise_{ts}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        logger.error(f"/api/export error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

def find_free_port(start_port=5000, max_port=5020):
    for port in range(start_port, max_port):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            if s.connect_ex(('127.0.0.1', port)) != 0: return port
    return start_port

def open_browser(port):
    webbrowser.open_new(f"http://127.0.0.1:{port}/")

if __name__ == "__main__":
    Thread(target=monitor_heartbeat, daemon=True).start()
    available_port = find_free_port()
    Timer(1.5, lambda: open_browser(available_port)).start()
    
    try:
        from waitress import serve
        logger.info(f"Starting Waitress production server on port {available_port}")
        serve(app, host="127.0.0.1", port=available_port, threads=8)
    except ImportError:
        logger.warning("Waitress not found. Falling back to Flask single-threaded development server.")
        app.run(debug=False, port=available_port)